from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
from pydantic import BaseModel
import json
import os
from datetime import datetime
from pathlib import Path
import asyncio
from contextlib import asynccontextmanager

# Import database and email modules
from database import (
    get_today_attendance,
    get_all_attendance,
    get_attendance_by_date,
    get_statistics,
    get_present_students,
    get_absent_students,
    get_all_students,
    clear_all_attendance
)

from email_scheduler import (
    send_daily_report,
    send_test_email,
    load_config as load_email_config,
    save_config as save_email_config,
    is_configured as is_email_configured,
    email_scheduler_task
)

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: Start background tasks
    asyncio.create_task(watch_database())
    asyncio.create_task(email_scheduler_task())  # Start email scheduler
    yield
    # Shutdown logic if needed

app = FastAPI(title="Face Recognition Attendance System", lifespan=lifespan)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()

# Paths
STATIC_DIR = Path("static")

# Create static directory if it doesn't exist
STATIC_DIR.mkdir(exist_ok=True)



# ============================================================================
# Chat Agent & Email Logic
# ============================================================================

class ChatMessage(BaseModel):
    message: str

class EmailConfig(BaseModel):
    recipient_email: Optional[str] = None
    sender_email: Optional[str] = None
    app_password: Optional[str] = None
    send_time: Optional[str] = None
    enabled: Optional[bool] = None



# ============================================================================
# Chat Agent
# ============================================================================

def process_agent_query(query: str):
    """Simple Rule-based AI Agent to answer questions about attendance"""
    query = query.lower()
    all_students_list = get_all_students()
    today_attendance = get_today_attendance()
    present_students = get_present_students()
    absent_students = get_absent_students()
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    # 1. Greeting
    if any(x in query for x in ['hi', 'hello', 'hey', 'who are you']):
        return "Hello! I am DOPA, your AI Attendance Assistant. Ask me about today's attendance or past records."

    # 2. "Who is present today?"
    if 'who' in query and 'present' in query and ('today' in query or 'now' in query):
        if not present_students:
            return "No one is present yet today."
        return f"Today ({today_date}), the following people are present: {', '.join(present_students)}."
    
    # 3. "Who is absent today?"
    if 'who' in query and 'absent' in query and ('today' in query or 'now' in query):
        if not absent_students:
            return "Everyone is present today!"
        return f"Today ({today_date}), the following people are absent: {', '.join(absent_students)}."

    # 4. "How many people?"
    if 'how many' in query:
        return f"Total {len(present_students)} people are marked present today, {len(absent_students)} are absent."

    # 5. Check specific person
    for name in all_students_list:
        if name.lower() in query:
            if 'today' in query or 'now' in query:
                if name in present_students:
                    # Get time
                    person_record = next((r for r in today_attendance if r['Name'] == name), None)
                    if person_record:
                        return f"Yes, {name} is present. Arrived at {person_record['Time']}."
                    else:
                        return f"Yes, {name} is present."
                else:
                    return f"No, {name} is absent today."
            
            # General stats for person
            all_records = get_all_attendance()
            total_days = len([r for r in all_records if r['Name'] == name])
            return f"{name} has been present for a total of {total_days} days recorded in the system."

    # 6. Email / Report
    if 'email' in query or 'send' in query or 'report' in query:
        return "To send a report, please use the 'Send Daily Report' button in the dashboard."

    return "I'm not sure I understand. Try asking: 'Who is present?', 'Who is absent?', 'Is [Name] here?', or 'How many people today?'"

# ============================================================================
# API Endpoints
# ============================================================================

@app.post("/api/chat")
async def chat_agent(chat_msg: ChatMessage):
    """Chat with the AI Agent"""
    response = process_agent_query(chat_msg.message)
    return {"response": response}

# ============================================================================
# Email Endpoints
# ============================================================================

@app.post("/api/email/send")
async def send_email_now(config: EmailConfig):
    """Send daily summary email now"""
    recipient = config.recipient_email if config.recipient_email else load_email_config().get('teacher_email')
    success, msg = send_daily_report(recipient_email=recipient)
    return {"success": success, "message": msg}

@app.post("/api/email/test")
async def test_email():
    """Send test email"""
    success, msg = send_test_email()
    return {"success": success, "message": msg}

@app.get("/api/email/config")
async def get_email_config():
    """Get email configuration (without password)"""
    config = load_email_config()
    # Don't send password to client
    safe_config = {
        "sender_email": config.get('sender_email', ''),
        "teacher_email": config.get('teacher_email', ''),
        "send_time": config.get('send_time', '17:00'),
        "enabled": config.get('enabled', False),
        "is_configured": is_email_configured()
    }
    return {"success": True, "data": safe_config}

@app.post("/api/email/config")
async def update_email_config(config: EmailConfig):
    """Update email configuration"""
    current_config = load_email_config()
    
    # Update only provided fields
    if config.sender_email is not None:
        current_config['sender_email'] = config.sender_email
    if config.app_password is not None:
        current_config['app_password'] = config.app_password
    if config.recipient_email is not None:
        current_config['teacher_email'] = config.recipient_email
    if config.send_time is not None:
        current_config['send_time'] = config.send_time
    if config.enabled is not None:
        current_config['enabled'] = config.enabled
    
    save_email_config(current_config)
    return {"success": True, "message": "Email configuration updated"}

# ============================================================================
# Attendance Endpoints
# ============================================================================

@app.get("/")
async def read_root():
    """Serve the main dashboard"""
    return FileResponse("static/index.html")

@app.get("/api/attendance/today")
async def get_today():
    """Get today's attendance"""
    today_data = get_today_attendance()
    return {
        "success": True,
        "data": today_data,
        "count": len(today_data)
    }

@app.get("/api/attendance/all")
async def get_all():
    """Get all attendance records"""
    data = get_all_attendance()
    return {
        "success": True,
        "data": data,
        "count": len(data)
    }

@app.get("/api/statistics")
async def get_stats():
    """Get attendance statistics"""
    stats = get_statistics()
    return {
        "success": True,
        "data": stats
    }

@app.get("/api/attendance/by-date/{date}")
async def get_by_date(date: str):
    """Get attendance by specific date (YYYY-MM-DD)"""
    filtered = get_attendance_by_date(date)
    return {
        "success": True,
        "date": date,
        "data": filtered,
        "count": len(filtered)
    }

@app.get("/api/students/present/{date}")
async def get_present(date: str = None):
    """Get present students for a date"""
    if date == 'today' or date is None:
        date = None  # Will use today
    present = get_present_students(date)
    return {
        "success": True,
        "data": present,
        "count": len(present)
    }

@app.get("/api/students/absent/{date}")
async def get_absent(date: str = None):
    """Get absent students for a date"""
    if date == 'today' or date is None:
        date = None  # Will use today
    absent = get_absent_students(date)
    return {
        "success": True,
        "data": absent,
        "count": len(absent)
    }

@app.get("/api/students/all")
async def get_students():
    """Get all registered students"""
    students = get_all_students()
    return {
        "success": True,
        "data": students,
        "count": len(students)
    }

@app.get("/api/export/csv/{date}")
async def export_csv(date: str):
    """Export attendance for a specific date as Excel with enhanced formatting"""
    # Filter for date (or 'all')
    if date == 'all':
        filtered = get_all_attendance()
        filename = f"attendance_report_all.xlsx"
        title_date = "All Records"
    else:
        filtered = get_attendance_by_date(date)
        filename = f"attendance_report_{date}.xlsx"
        title_date = date
    
    # Get all students for comprehensive report
    all_students = get_all_students()
    present_students = [r['Name'] for r in filtered]
    
    # Create Excel workbook
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Title
        ws['A1'] = f"Attendance Report - {title_date}"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Summary
        row = 3
        ws[f'A{row}'] = "Summary:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = f"Total Present: {len(present_students)}"
        ws[f'B{row}'] = f"Total Absent: {len(all_students) - len(present_students)}"
        ws[f'C{row}'] = f"Total Students: {len(all_students)}"
        row += 2
        
        # Present Students Table
        ws[f'A{row}'] = "Present Students"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Headers for present students
        headers = ['Name', 'Date', 'Time', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data rows for present students
        for record in filtered:
            ws.cell(row=row, column=1, value=record.get('Name', ''))
            ws.cell(row=row, column=2, value=record.get('Date', ''))
            ws.cell(row=row, column=3, value=record.get('Time', ''))
            ws.cell(row=row, column=4, value='Present')
            
            # Green fill for present
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            row += 1
        
        row += 1
        
        # Absent Students Table
        absent_students = [s for s in all_students if s not in present_students]
        if absent_students:
            ws[f'A{row}'] = "Absent Students"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Headers for absent students
            headers_absent = ['Name', 'Date', 'Status', '']
            for col, header in enumerate(headers_absent, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="F87171", end_color="F87171", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data rows for absent students
            for name in absent_students:
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=date if date != 'all' else 'N/A')
                ws.cell(row=row, column=3, value='Absent')
                
                # Red fill for absent
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to bytes
        import io
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        # Fallback to CSV if openpyxl not available
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Name', 'Date', 'Time', 'Status'])
        
        # Rows
        for record in filtered:
            writer.writerow([
                record.get('Name', ''),
                record.get('Date', ''),
                record.get('Time', ''),
                record.get('Status', 'Present')
            ])
            
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8')),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=attendance_report_{date}.csv"}
        )

@app.post("/api/attendance/clear")
async def clear_attendance():
    """Clear all attendance records"""
    clear_all_attendance()
    await manager.broadcast({
        "type": "attendance_cleared",
        "message": "All attendance records cleared"
    })
    return {"success": True, "message": "Attendance cleared"}

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time updates"""
    await manager.connect(websocket)
    try:
        # Send initial data
        await websocket.send_json({
            "type": "initial_data",
            "attendance": get_today_attendance(),
            "statistics": get_statistics()
        })
        
        # Keep connection alive and listen for messages
        while True:
            data = await websocket.receive_text()
            # Echo back or handle commands
            await websocket.send_json({
                "type": "pong",
                "message": "Connection alive"
            })
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# ============================================================================
# File Watcher (monitors database for changes)
# ============================================================================

async def watch_database():
    """Watch database for changes and broadcast updates"""
    import time
    last_count = 0
    
    while True:
        try:
            current_data = get_today_attendance()
            current_count = len(current_data)
            
            if current_count != last_count:
                last_count = current_count
                # Broadcast update to all connected clients
                await manager.broadcast({
                    "type": "attendance_update",
                    "attendance": current_data,
                    "statistics": get_statistics()
                })
        except Exception as e:
            print(f"Error watching database: {e}")
        
        await asyncio.sleep(2)  # Check every 2 seconds



# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

if __name__ == "__main__":
    import uvicorn
    print("="*70)
    print("Starting Face Recognition Attendance Web Server")
    print("="*70)
    print("Dashboard: http://localhost:8000")
    print("API Docs: http://localhost:8000/docs")
    print("="*70)
    uvicorn.run(app, host="0.0.0.0", port=8000)
